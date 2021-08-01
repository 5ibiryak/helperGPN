from PyQt5 import QtCore, QtGui, QtWidgets
import sys
from project_4 import Ui_MainWindow
from tkinter import *
import os
from openpyxl import load_workbook
import pandas as pd
 
dirlist = 'D:/'
myFileName = ""

def btn_click():
    name = ui.textEdit_object.toPlainText()
    date = ui.dateEdit.dateTime().toString('dd.MM.yyyy')
    dirlist = ui.plainTextEdit_folder.toPlainText()
    os.chdir(dirlist)
    #создание главной папки
    os.mkdir(name + '_' + date)
    os.chdir(name + '_' + date)
    #создание подпапок в главной
    os.mkdir("Field")
    os.mkdir("P4D_processing")
    os.mkdir('Results' + '_' + name + '_' + date)
    #создание папок в Results
    os.chdir('Results' + '_' + name + '_' + date)
    os.mkdir('DEM')
    os.mkdir('Ortho')
    os.mkdir('Point_cloud')
    print('Okey')
    
    # # добавление записи о полете в таблицу
    # os.chdir(dirlist)
    # wb = load_workbook(myFileName)
    # ws = wb.worksheets[0]
    # name = name+''
    # date = date+''
    # # проверка не пуста ли переменная имя
    # if name:
    #     newItemID = ws.max_row
    #     ws.append([newItemID, name, date])
    # else:
    #     print('No name entered')
    # #завершение работы с таблицой
    # wb.save(filename=myFileName)
    # wb.close()
    
# Работа с таблицей
    global myFileName
    arr = myFileName.split("/")
    filedir=""
    i = 0
    while i < (len(arr)-1):
        filedir = filedir + arr[i] +'/'
        i = i+1
    myFileName = arr[len(arr)-1]
    
    os.chdir(filedir)
    wb = load_workbook(myFileName)
    
    #проверка сушествует ли таблица в выбранном файле
    df = pd.read_excel(myFileName)
    check = df.empty
    # создание таблицы, если ее нет
    if (check):
        ws = wb.worksheets[0]
        ws.append(['Порядковый номер', 'Наименование объекта','Наименование проекта', 'Дата съемки'])
    
    # добавление записи в таблицу
    ws = wb.worksheets[0]
    name = name+''
    date = date+''
    project_name = name+'_'+date
    # проверка не пуста ли переменная имя
    if name:
        newItemID = ws.max_row
        ws.append([newItemID, name,project_name, date])
    else:
        print('No name entered')
    #завершение работы с таблицой
    wb.save(filename=myFileName)
    wb.close()



#выбор папки
def getDirectory():
    dirlist = QtWidgets.QFileDialog.getExistingDirectory()
    ui.plainTextEdit_folder.setPlainText(format(dirlist))
    ui.plainTextEdit_folder.setGeometry(QtCore.QRect(240, 90, 321, 46))

#выбор файла
def getFileName():
        filename = QtWidgets.QFileDialog.getOpenFileName()
        ui.plainTextEdit_excel.appendHtml(format(filename))
        ui.plainTextEdit_excel.setGeometry(QtCore.QRect(240, 180, 321, 46))

app = QtWidgets.QApplication(sys.argv)
MainWindow = QtWidgets.QMainWindow()
ui = Ui_MainWindow()
ui.setupUi(MainWindow)
MainWindow.show()

ui.plainTextEdit_folder.appendHtml(format(dirlist))

ui.pushButton_folder.clicked.connect(getDirectory)
ui.pushButton_excel.clicked.connect(getFileName)
ui.pushButton.clicked.connect(btn_click)

sys.exit(app.exec_())
