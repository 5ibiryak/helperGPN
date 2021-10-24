from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox
import sys
from design_app import *
from tkinter import *
import os
from openpyxl import load_workbook
import pandas as pd




dirlist = 'C:/'
myFileName = ""
myLogText = "" # строка для логов5
projID = 0 # id проекта в таблице


### путь для приложений
with open('path.txt',"r") as myfile:
  data = myfile.readlines()

str1=data[0]
str2=data[1]
str3=data[2]
path_Metashape = str1[:-1]
path_ContexCapture = str2[:-1]
path_Pix4D = str3[:-1]


def btn_click():
    
    # СЧИТЫВАНИЕ ДАННЫХ
    name = ui.textEdit_object.toPlainText()
    date = ui.dateEdit.dateTime().toString('ddMMyyyy')
    dirlist = ui.plainTextEdit_folder.toPlainText()
    os.chdir(dirlist)
    
    
    # проверка что имя введено
    if not name:
        myLogText = 'Введите имя проекта'
        showErrMsg(myLogText)
    else:
        # РАБОТА С ТАБЛИЦЕЙ
        global myFileName
        arr = myFileName.split("/")
        i = 0
        filedir = ""
        while i < (len(arr)-1):
            filedir = filedir + arr[i] +'/'
            i = i+1
        print(myFileName,filedir)
        filename = arr[len(arr)-1]
        os.chdir(filedir)
        
        # проверка закрыт ли файл с таблицей
        try:
            myfile = open(filename, "r+")
        except IOError:
            myLogText = f'Файл {filename} открыт. Для продолжения закройте файл и нажмите кнопку "Создать" еще раз.'
            showErrMsg(myLogText)
        else:    
            wb = load_workbook(filename)
            #проверка сушествует ли таблица в выбранном файле
            df = pd.read_excel(filename)
            check = df.empty
            # создание таблицы, если ее нет
            if (check):
                ws = wb.worksheets[0]
                ws.append(['Порядковый номер', 'Наименование объекта','Наименование проекта', 'Дата съемки'])
                myLogText = 'Новая таблица успешно создана в файле '+filename
                showInfoMsg(myLogText)
                        
            ws = wb.worksheets[0]
            name = name+''
            date = date+''
            newItemID = ws.max_row
            
             
            # проверка параметра чекбокса (добавлять id к наименованию проекта - true или нет - false)
            if ui.checkBox.isChecked():
                if newItemID == 1:
                    projID = 1
                else:
                    # получаем порядковый номер последнего проекта в таблице
                    found = False
                    while not found:
                        mstring = f'A{newItemID}'
                        projID = ws[mstring].value
                        if projID == None:
                            newItemID = newItemID - 1
                        else:
                            found = True
                    mstring = f'A{newItemID}'
                    projID = ws[mstring].value + 1
                project_name = str(projID)+'_'+name+'_'+date
            else:
                project_name = name+'_'+date
            
            # проверка что заданной директории не существует
            if os.path.isdir(project_name):
                myLogText = 'Папка с заданным именем уже существует'
                showErrMsg(myLogText)
            else:
                # добавление записи в таблицу
                if ui.checkBox.isChecked():
                    ws.append([projID, name,project_name, date])
                else:
                    ws.append(['', name,project_name, date])
                    
                #завершение работы с таблицeй
                wb.save(filename=filename)
                wb.close()
                myLogText = f'Запись о проекте {project_name} успешно добавлена в таблицу'
                showInfoMsg(myLogText)
                    
                # СОЗДАНИЕ ПАПОК
                #создание главной папки
                os.mkdir(project_name)
                os.chdir(project_name)
                #создание подпапок в главной
                os.mkdir("Field")
                use_program = ui.comboBox.currentText()
                if use_program == "Agisoft Metashape":
                    use_program = "MS_processing"
                elif use_program == "Contex capture":
                    use_program = "CC_processing"
                elif use_program == "Pix4D":
                    use_program = "P4D_processing"
                os.mkdir(use_program)
                os.mkdir('Results' + '_' + name + '_' + date)
                #создание папок в Results
                os.chdir('Results' + '_' + name + '_' + date)
                os.mkdir('DEM')
                os.mkdir('Ortho')
                os.mkdir('Point_cloud')
                print('Okey')
                myLogText = f'Система хранения для проекта {project_name} успешно создана'
                showInfoMsg(myLogText)
                
                #открытие программы
                window_for_open_app()
                
 #Открытие приложения MS,Px4 после создания документа
def window_for_open_app ():
    # def msgButtonClick(i):
    #     print("Button clicked is:",i.text())
    app_msg = QMessageBox()
    app_msg.setIcon(QMessageBox.Information)
    app_msg.setText("Открыть программу?")
    app_msg.setWindowTitle("Запуск программы")
    app_msg.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
    #app_msg.buttonClicked.connect(msgButtonClick)
    returnValue = app_msg.exec()
    if returnValue == QMessageBox.Ok:
        print('OK clicked')
        try:           
            os.startfile(path_Metashape)
        except:
            pass
    else:
        sys.exit()

#выбор папки
def getDirectory():
    dirlist = QtWidgets.QFileDialog.getExistingDirectory()
    ui.plainTextEdit_folder.setPlainText(format(dirlist))
    ui.plainTextEdit_folder.setGeometry(QtCore.QRect(240, 90, 321, 46))

#выбор файла
def getFileName():
    filename,_ = QtWidgets.QFileDialog.getOpenFileName()
    global myFileName
    myFileName = filename
    ui.plainTextEdit_excel.appendHtml(format(filename))
    ui.plainTextEdit_excel.setGeometry(QtCore.QRect(240, 180, 321, 46))

app = QtWidgets.QApplication(sys.argv)
MainWindow = QtWidgets.QMainWindow()

# создание окошка для сообшений об ошибках
err_msg = QMessageBox()
err_msg.setIcon(QMessageBox.Critical)
err_msg.setInformativeText('')
err_msg.setWindowTitle("Ошибка")

def showErrMsg(mstring):
    err_msg.setInformativeText(mstring)
    err_msg.exec_()

# создание окошка для инфо-сообщений
info_msg = QMessageBox()
info_msg.setIcon(QMessageBox.Information)
info_msg.setInformativeText('')
info_msg.setWindowTitle("Ход работы")

def showInfoMsg(mstring):
    info_msg.setInformativeText(mstring)
    info_msg.exec_()



###########запуск приложения##############
ui = Ui_MainWindow()
ui.setupUi(MainWindow)
MainWindow.show()

ui.plainTextEdit_folder.appendHtml(format(dirlist))

ui.pushButton_folder.clicked.connect(getDirectory)
ui.pushButton_excel.clicked.connect(getFileName)
ui.pushButton.clicked.connect(btn_click)


sys.exit(app.exec_())
